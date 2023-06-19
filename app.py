from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# create new excel obj
# wb = Workbook()


wb = load_workbook('test.xlsx')
ws = wb.active

# for row in range(1,11):
#     for col in range(1, 5):
#         char = get_column_letter(col)
#         ws[char + str(row)].value = char + str(row)

# ws.merge_cells("A1:D1")
# ws.unmerge_cells("A1:D1")

# ws.insert_rows(7)

# ws.insert_cols(2)

ws.move_range("C1:D11", rows=2, cols=2)

wb.save('test.xlsx')