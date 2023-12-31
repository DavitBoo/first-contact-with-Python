from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from datetime import datetime

from find_kw import get_kw, get_sentence, get_links, check_link_status
import requests

from urllib.parse import urlparse, urljoin


def validate_url(url):
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}
        response = requests.head(url, headers=headers)

        if response.status_code == 200:
            return True
        else:
            print(
                f"La URL {url} no es accesible. Verifica la URL e intenta nuevamente.")
    except requests.exceptions.RequestException as e:
        print(f"Error al acceder a la URL {url}: {str(e)}")

    return False


def process_url(url):
    parsed_url = urlparse(url)
    if not parsed_url.scheme:
        url = 'https://' + url
    return url


url = input(
    'Inseta la URL para analizar las palabras clave. ej: www.google.com:\n')
query_url = process_url(url)

if validate_url(query_url):

    wb = Workbook()
    keyword_sheet = wb.active
    keyword_sheet.title = "Keywords"

    headings = ['Keyword', 'Times Repeated']
    keyword_sheet.append(headings)

    word_list = get_kw(query_url)
    print(word_list)

    coords = 1
    for word, count in word_list:
        print(f"{word}: {count}")
        coords += 1
        keyword_sheet.append([word, count])

    sentence_list = get_sentence(query_url)

    sentence_sheet = wb.create_sheet(title='Sentences')
    sentence_sheet.append(['Sentence', 'Times Repeated'])

    for phrase, count in sentence_list:
        phrase_str = " ".join(phrase)
        sentence_sheet.append([phrase_str, count])

    internal_links, external_links = get_links(query_url)
    checked_urls = set()

    internal_links_sheet = wb.create_sheet(title='Internal Links')
    internal_links_sheet.append(['Link', 'Status'])

    for link in internal_links:
        link_url = urljoin(query_url, link)
        if link_url not in checked_urls:
            status = check_link_status(link_url)
            internal_links_sheet.append([link, status])
            checked_urls.add(link_url)

    external_links_sheet = wb.create_sheet(title='External Links')
    external_links_sheet.append(['Link', 'Status'])

    for link in external_links:
        link_url = urljoin(query_url, link)
        if link_url not in checked_urls:
            status = check_link_status(link_url)
            external_links_sheet.append([link, status])
            checked_urls.add(link_url)

    current_timestamp = datetime.now()
    formatted_timestamp = current_timestamp.strftime('%Y%m%d%H%M%S')

    for col in range(1, 3):
        keyword_sheet[get_column_letter(
            col) + '1'].font = Font(bold=True, color="0066AAFF")
        sentence_sheet[get_column_letter(
            col) + '1'].font = Font(bold=True, color="0066AAFF")
        internal_links_sheet[get_column_letter(
            col) + '1'].font = Font(bold=True, color="0066AAFF")
        external_links_sheet[get_column_letter(
            col) + '1'].font = Font(bold=True, color="0066AAFF")

    # eliminamos http o https para crear un nombre de archivo amigable al usuario
    if "http://" in url:
        url = url.replace("http://", "")
    elif "https://" in url:
        url = url.replace("https://", "")

    if "/" in url:
        url = url.split("/", 1)[0]

    wb.save(f'{formatted_timestamp}-{url}.xlsx')
else:
    print("La URL proporcionada no es válida o no se puede acceder a ella. Verifica la URL e intenta nuevamente.")
